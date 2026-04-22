"""
q_report worker — claude_verdict.json을 읽어 PDF + XLSX 리포트 생성, NFS 저장, DB 업데이트.
concurrency=2 (celeryconfig).
"""

import asyncio
import json
import shutil
import subprocess
import tempfile
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import structlog
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from config.settings import settings
from workers.app import app
from workers.notify import publish_job_status

log = structlog.get_logger(__name__)

# 리포트는 KST(UTC+9) 기준으로 시간 표시
KST = timezone(timedelta(hours=9))


def _make_session() -> tuple:
    """매 asyncio.run() 루프마다 새 엔진+세션팩토리를 생성."""
    engine = create_async_engine(settings.database_url, echo=False, pool_pre_ping=True)
    session_local = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    return engine, session_local


_TEMPLATES_DIR = Path(__file__).parent.parent / "templates"
_PROFILES_DIR = Path(__file__).parent.parent / "checks" / "profiles"

# 스크립트명 → 리포트에 표시할 한글 이름.
# 없으면 스크립트명 그대로 폴백.
DISPLAY_NAMES: dict[str, str] = {
    "sw_gpu_hw": "GPU 하드웨어 (PCIe)",
    "sw_cpu": "CPU 정보·온도",
    "sw_memory": "메모리 (DIMM/NUMA/ECC)",
    "sw_storage_hw": "스토리지 하드웨어",
    "sw_network": "네트워크 (NIC/IB)",
    "sw_os_version": "OS·커널 버전",
    "sw_power_mgmt": "전원 관리",
    "sw_auto_update": "자동 업데이트 비활성화",
    "sw_gpu_sw": "GPU 드라이버·CUDA",
    "sw_storage_sw": "스토리지 SW (NVMe SMART)",
    "stress_gpu": "GPU 부하 테스트",
    "stress_cpu": "CPU 부하 테스트",
    "nccl_bandwidth": "NCCL 대역폭",
    "collect_all_logs": "로그 수집",
}

# 스크립트별로 리포트에 표시할 필드 목록.
# 없는 스크립트는 parse_detail() 결과 전체를 최대 8항목 fallback 표시.
KNOWN_FIELDS: dict[str, list[str]] = {
    "sw_gpu_hw": [
        "gpu_count",
        "link_width",
        "link_speed",
        "link_state_width",
        "link_state_speed",
        "downgraded_gpus",
    ],
    "sw_cpu": ["model", "sockets", "cores", "max_temp_c"],
    "sw_memory": ["total_gb", "dimm_count", "numa_nodes", "memory_errors"],
    "sw_storage_hw": ["disk_count", "devices", "md_degraded"],
    "sw_network": ["active_nics", "up_ifaces", "ib_active"],
    "sw_os_version": ["name", "version_id", "kernel", "supported_os"],
    "sw_power_mgmt": ["sleep_target", "cpu_governor", "max_cstate"],
    "sw_auto_update": ["unattended_active"],
    "sw_gpu_sw": [
        "gpu_count",
        "driver",
        "cuda",
        "vram_gb",
        "gpu_max_temp_c",
        "ecc_delta_uncorr",
        "persistence",
    ],
    "sw_storage_sw": ["root_used_pct", "nvme_count", "nvme_critical", "nvme_wear_high"],
    "stress_gpu": [
        "tool",
        "duration_s",
        "gpu_class",
        "peak_temp_c",
        "peak_power_w",
        "power_ratio_pct",
        "avg_util_pct",
        "ecc_delta_uncorr",
        "cooling_consistency_score",
        "cooling_avg_spread_c",
        "cooling_max_spread_c",
        "cooling_outlier_gpus",
    ],
    "stress_cpu": [
        "tool",
        "duration_s",
        "peak_temp_c",
        "max_freq_mhz",
        "avg_util_pct",
        "throttle_sample_count",
    ],
    "nccl_bandwidth": [
        "tool",
        "gpu_count",
        "bw_2gpu_gbs",
        "bw_4gpu_gbs",
        "min_bw_2gpu_gbs",
        "min_bw_4gpu_gbs",
    ],
    "collect_all_logs": ["xid_count"],
}


def parse_detail(detail: str) -> tuple[dict[str, str], list[str]]:
    """'key=val|key2=val2|WARN:msg' 형식의 detail → (data_dict, diagnostics_list).

    WARN:/FAIL:/INFO: 프리픽스 메시지는 diagnostics_list에 별도 수집.
    일반 key=val은 dict에 담김.
    """
    if not detail:
        return {}, []
    data: dict[str, str] = {}
    diagnostics: list[str] = []
    for part in str(detail).split("|"):
        part = part.strip()
        if not part:
            continue
        if part.startswith(("WARN:", "FAIL:", "INFO:")):
            diagnostics.append(part)
            continue
        k, sep, v = part.partition("=")
        if sep:
            data[k.strip()] = v.strip()
    return data, diagnostics


def _build_phase_map(product_profile: str) -> dict[str, str]:
    """profile JSON에서 {script_name: phase_name} 역방향 매핑 생성.

    profile 파일이 없거나 파싱 실패 시 빈 dict 반환 → 모든 결과가 'unknown' 으로 분류됨.
    """
    profile_path = _PROFILES_DIR / f"{product_profile}.json"
    if not profile_path.exists():
        return {}
    try:
        profile = json.loads(profile_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    mapping: dict[str, str] = {}
    for phase_name, phase_data in profile.get("phases", {}).items():
        for script in phase_data.get("scripts", []):
            mapping[script] = phase_name
    return mapping


def _enrich_results(raw_results: list[dict], script_to_phase: dict[str, str]) -> list[dict]:
    """check_results에 phase, display_fields 추가.

    display_fields 구성 순서:
    1. 진단 메시지 (WARN:/FAIL:/INFO:) — 가장 눈에 띄도록 맨 앞
    2. KNOWN_FIELDS에 정의된 필드 순서대로 (없으면 parsed 상위 8개)
    """
    enriched = []
    for cr in raw_results:
        parsed, diagnostics = parse_detail(cr.get("detail") or "")
        phase = script_to_phase.get(cr["check_name"], "unknown")
        known = KNOWN_FIELDS.get(cr["check_name"])
        if known is not None:
            fields = [(k, parsed[k]) for k in known if k in parsed]
            if not fields and not diagnostics:
                fields = list(parsed.items())[:8]
        else:
            fields = list(parsed.items())[:8]
        # 진단 메시지를 (label, body) 튜플로 display_fields 맨 앞에 추가
        # 예: "WARN:no_stress_tool_temp_only" → ("WARN", "no_stress_tool_temp_only")
        diag_fields = []
        for msg in diagnostics:
            label, _, body = msg.partition(":")
            diag_fields.append((label, body or msg))
        display_fields = diag_fields + fields
        display_name = DISPLAY_NAMES.get(cr["check_name"], cr["check_name"])
        enriched.append(
            {
                **cr,
                "phase": phase,
                "display_name": display_name,
                "display_fields": display_fields,
            }
        )
    return enriched


def _item_to_str(it: dict) -> str:
    """fail_items/warn_items dict → 사람이 읽을 수 있는 문자열."""
    base = f"{it.get('check', '?')}: {it.get('metric', '?')}={it.get('value', '?')}"
    rule = it.get("rule")
    threshold = it.get("threshold")
    if rule and threshold is not None:
        return f"{base} ({rule}={threshold})"
    return base


# LaTeX 전용 Jinja2 환경 — \BLOCK{}, \VAR{} 구문 사용 (LaTeX {} 충돌 방지)
def _latex_escape(text: str) -> str:
    """LaTeX 특수문자 이스케이프."""
    text = str(text)
    for old, new in [
        ("\\", r"\textbackslash{}"),
        ("&", r"\&"),
        ("%", r"\%"),
        ("$", r"\$"),
        ("#", r"\#"),
        ("_", r"\_"),
        ("{", r"\{"),
        ("}", r"\}"),
        ("~", r"\textasciitilde{}"),
        ("^", r"\textasciicircum{}"),
    ]:
        text = text.replace(old, new)
    return text


_latex_env = Environment(
    block_start_string=r"\BLOCK{",
    block_end_string="}",
    variable_start_string=r"\VAR{",
    variable_end_string="}",
    comment_start_string=r"\#{",
    comment_end_string="}",
    trim_blocks=True,
    lstrip_blocks=True,
    autoescape=False,
    loader=FileSystemLoader(str(_TEMPLATES_DIR)),
)


def _detail_latex(text: str) -> str:
    """detail 필드(key=val|key2=val2)를 LaTeX p{} 컬럼에서 줄바꿈되도록 변환.

    | 구분자를 기준으로 분리 후 각 파트를 latex_escape → \\newline 으로 재결합.
    """
    parts = str(text).split("|")
    return r"\newline ".join(_latex_escape(p.strip()) for p in parts)


def _fields_latex(display_fields: list) -> str:
    """(key, val) 튜플 목록 → LaTeX \\newline 구분 이스케이프 문자열.

    WARN/FAIL/INFO 레이블은 색상 강조. 그 외는 `key=val` 형식.
    """
    if not display_fields:
        return "—"
    _DIAG_COLOR = {"WARN": "warntext", "FAIL": "failtext", "INFO": "headerblue"}
    parts = []
    for k, v in display_fields:
        color = _DIAG_COLOR.get(k)
        if color:
            # 진단 메시지: 색상 + bold 레이블 + 본문
            parts.append(rf"\textcolor{{{color}}}{{\bfseries {k}}}: {_latex_escape(v)}")
        else:
            parts.append(_latex_escape(f"{k}={v}"))
    return r"\newline ".join(parts)


_latex_env.filters["latex_escape"] = _latex_escape
_latex_env.filters["detail_latex"] = _detail_latex
_latex_env.filters["fields_latex"] = _fields_latex


# ---------------------------------------------------------------------------
# PDF 생성
# ---------------------------------------------------------------------------


def _render_pdf(context: dict, output_path: Path) -> None:
    """LaTeX → xelatex 컴파일 → PDF."""
    tex_str = _latex_env.get_template("report.tex.j2").render(**context)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        tex_file = tmpdir_path / "report.tex"
        tex_file.write_text(tex_str, encoding="utf-8")

        result = subprocess.run(
            [
                "xelatex",
                "-interaction=nonstopmode",
                "-halt-on-error",
                f"-output-directory={tmpdir}",
                str(tex_file),
            ],
            capture_output=True,
            text=True,
            timeout=120,
            cwd=tmpdir,
        )
        if result.returncode != 0:
            raise RuntimeError(f"xelatex failed:\n{result.stdout[-3000:]}")

        shutil.copy(tmpdir_path / "report.pdf", output_path)


# ---------------------------------------------------------------------------
# XLSX 생성
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PASS_FONT = Font(bold=True, color="155724")
_FAIL_FONT = Font(bold=True, color="721C24")
_WARN_FONT = Font(bold=True, color="856404")

_STATUS_FONT = {"pass": _PASS_FONT, "fail": _FAIL_FONT, "warn": _WARN_FONT}


def _render_xlsx(context: dict, output_path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── 시트1: 요약 ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "요약"

    summary_rows = [
        ("Job ID", context["job_id"]),
        ("대상 서버", context["target_host"]),
        ("접속 유저", context["target_user"]),
        ("제품 프로파일", context["product_profile"]),
        ("검수 시작", context["created_at"]),
        ("리포트 생성", context["generated_at"]),
        ("최종 판정", context["overall"].upper()),
    ]
    for row in summary_rows:
        ws_summary.append(row)
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)

    verdict_cell = ws_summary.cell(row=7, column=2)
    verdict_cell.font = _STATUS_FONT.get(context["overall"], Font(bold=True))

    ws_summary.append([])
    if context["fail_reasons"]:
        ws_summary.append(["FAIL 사유"])
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, color="721C24")
        for r in context["fail_reasons"]:
            ws_summary.append(["", r])

    if context["warn_reasons"]:
        ws_summary.append(["주의 사항"])
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, color="856404")
        for r in context["warn_reasons"]:
            ws_summary.append(["", r])

    if context["summary"]:
        ws_summary.append([])
        ws_summary.append(["Claude 요약"])
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
        ws_summary.append(["", context["summary"]])

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 60

    # ── 시트2: 검수 항목 상세 ───────────────────────────────────────────────
    ws_detail = wb.create_sheet("검수 상세")
    headers = ["검수 항목", "스크립트", "Phase", "상태", "Claude 판정", "상세"]
    ws_detail.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for cr in context["check_results"]:
        ws_detail.append(
            [
                cr.get("display_name", cr["check_name"]),
                cr["check_name"],
                cr.get("phase", ""),
                cr["status"].upper(),
                cr["claude_verdict"] or "",
                cr["detail"],
            ]
        )
        status_cell = ws_detail.cell(row=ws_detail.max_row, column=4)
        status_cell.font = _STATUS_FONT.get(cr["status"], Font())
        status_cell.alignment = Alignment(horizontal="center")
        ws_detail.cell(row=ws_detail.max_row, column=6).alignment = Alignment(wrap_text=True)

    ws_detail.column_dimensions["A"].width = 26  # 검수 항목 (한글 display_name)
    ws_detail.column_dimensions["B"].width = 20  # 스크립트 (원본 파일명)
    ws_detail.column_dimensions["C"].width = 14  # Phase
    ws_detail.column_dimensions["D"].width = 8  # 상태
    ws_detail.column_dimensions["E"].width = 35  # Claude 판정
    ws_detail.column_dimensions["F"].width = 50  # 상세

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# DB 헬퍼
# ---------------------------------------------------------------------------


async def _load_job_and_results(session: AsyncSession, job_id: str) -> tuple:
    from api.models import CheckResult, Job

    job_result = await session.execute(select(Job).where(Job.id == uuid.UUID(job_id)))
    job = job_result.scalar_one_or_none()
    if job is None:
        raise ValueError(f"Job {job_id} not found")

    cr_result = await session.execute(
        select(CheckResult).where(CheckResult.job_id == uuid.UUID(job_id))
    )
    return job, list(cr_result.scalars().all())


async def _save_report_record(
    session: AsyncSession,
    job_id: str,
    pdf_path: str,
    xlsx_path: str,
) -> None:
    from api.models import Report

    report = Report(
        job_id=uuid.UUID(job_id),
        pdf_path=pdf_path,
        xlsx_path=xlsx_path,
    )
    session.add(report)
    await session.commit()


async def _update_job_status(
    session: AsyncSession,
    job_id: str,
    status: str,
    error_message: str | None = None,
) -> None:
    from api.models import Job

    result = await session.execute(select(Job).where(Job.id == uuid.UUID(job_id)))
    job = result.scalar_one_or_none()
    if job is None:
        raise ValueError(f"Job {job_id} not found")
    job.status = status
    job.updated_at = datetime.now(timezone.utc)
    if error_message:
        job.error_message = error_message[:2000]
    await session.commit()


# ---------------------------------------------------------------------------
# 차트 생성 (시계열 + 집계)
# ---------------------------------------------------------------------------


def _generate_all_charts(
    job_id: str,
    enriched_results: list[dict],
    product_profile: str,
    report_dir: Path,
) -> dict[str, str]:
    """검수 결과의 stress/nccl 차트를 PNG로 생성.

    반환: {"gpu": "/abs/path", "gpu_cooling": ..., "cpu": ..., "nccl": ...}
    매트플롯립 import 비용은 워커 프로세스당 1회.
    """
    from workers import report_charts

    raw_dir = Path(settings.nfs_base_path) / "results" / job_id / "inspect_raw"
    paths: dict[str, str] = {}

    # 사용자 온도 상한 (profile validation.rules에서 추출)
    user_temp_limit = _extract_user_temp_limit(product_profile)

    # GPU stress 차트
    gpu_ts = raw_dir / "stress_gpu_timeseries.json"
    gpu_chart = report_dir / "chart_stress_gpu.png"
    if gpu_ts.exists():
        try:
            ok = report_charts.generate_gpu_chart(gpu_ts, gpu_chart, user_temp_limit)
            if ok:
                paths["gpu"] = str(gpu_chart)
                cooling = report_dir / "chart_stress_gpu_cooling_heatmap.png"
                if cooling.exists():
                    paths["gpu_cooling"] = str(cooling)
        except Exception as exc:
            log.warning("report.chart_gpu_failed", job_id=job_id, error=str(exc))

    # CPU stress 차트
    cpu_ts = raw_dir / "stress_cpu_timeseries.json"
    cpu_chart = report_dir / "chart_stress_cpu.png"
    if cpu_ts.exists():
        try:
            if report_charts.generate_cpu_chart(cpu_ts, cpu_chart):
                paths["cpu"] = str(cpu_chart)
        except Exception as exc:
            log.warning("report.chart_cpu_failed", job_id=job_id, error=str(exc))

    # nccl bar chart (시계열 아님; enriched에서 detail 파싱)
    nccl_chart = report_dir / "chart_nccl.png"
    nccl_result = next((r for r in enriched_results if r["check_name"] == "nccl_bandwidth"), None)
    if nccl_result:
        parsed, _ = parse_detail(nccl_result.get("detail") or "")
        try:
            if report_charts.generate_nccl_chart(parsed, nccl_chart):
                paths["nccl"] = str(nccl_chart)
        except Exception as exc:
            log.warning("report.chart_nccl_failed", job_id=job_id, error=str(exc))

    log.info("report.charts_generated", job_id=job_id, charts=list(paths.keys()))
    return paths


def _extract_user_temp_limit(product_profile: str) -> float | None:
    """profile JSON의 validation.rules에서 sw_gpu_sw.gpu_max_temp_c.fail_above 값 추출."""
    profile_path = _PROFILES_DIR / f"{product_profile}.json"
    if not profile_path.exists():
        return None
    try:
        profile = json.loads(profile_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    for rule in profile.get("validation", {}).get("rules", []):
        if rule.get("check") == "sw_gpu_sw" and rule.get("metric") == "gpu_max_temp_c":
            val = rule.get("fail_above")
            if isinstance(val, (int, float)):
                return float(val)
    return None


# ---------------------------------------------------------------------------
# 핵심 async 로직
# ---------------------------------------------------------------------------


async def _async_generate_report(job_id: str) -> None:
    engine, SessionLocal = _make_session()
    try:
        # ── 1. DB 로드 ────────────────────────────────────────
        async with SessionLocal() as session:
            job, check_results = await _load_job_and_results(session, job_id)
            job_data = {
                "job_id": str(job.id),
                "target_host": job.target_host,
                "target_user": job.target_user,
                "product_profile": job.product_profile,
                "created_at": job.created_at.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S KST"),
            }

        # ── 2. NFS에서 Claude 판독 결과 로드 (없으면 DB에서 합성) ──
        verdict_file = Path(settings.nfs_base_path) / "results" / job_id / "claude_verdict.json"
        if verdict_file.exists():
            verdict = json.loads(verdict_file.read_text(encoding="utf-8"))
            overall = verdict.get("verdict", "fail")
            fail_reasons = [_item_to_str(it) for it in verdict.get("fail_items", [])]
            warn_reasons = [_item_to_str(it) for it in verdict.get("warn_items", [])]
            summary = (
                verdict.get("agent_verdict", {}).get("reason", "")
                if verdict.get("agent_verdict")
                else ""
            )
        else:
            # validation이 실행되지 않은 경우 (preflight/post_install 실패 등)
            # check_results에서 fallback verdict 합성
            log.warning("report.verdict_missing_fallback", job_id=job_id)
            fail_reasons = [
                f"{cr.check_name}: {cr.detail[:120]}" for cr in check_results if cr.status == "fail"
            ]
            warn_reasons = [
                f"{cr.check_name}: {cr.detail[:120]}" for cr in check_results if cr.status == "warn"
            ]
            overall = "fail" if fail_reasons else "warn" if warn_reasons else "fail"
            summary = ""

        # ── 3. 렌더링 컨텍스트 구성 ───────────────────────────
        # ORM 객체 → dict 변환 후 phase/display_fields 보강
        raw_results = [
            {
                "check_name": cr.check_name,
                "status": cr.status,
                "detail": cr.detail,
                "claude_verdict": cr.claude_verdict,
            }
            for cr in check_results
        ]
        script_to_phase = _build_phase_map(job_data["product_profile"])
        enriched = _enrich_results(raw_results, script_to_phase)

        def _by_phase(phase: str) -> list[dict]:
            return [r for r in enriched if r["phase"] == phase]

        all_statuses = [r["status"] for r in enriched]
        generated_at = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S KST")

        # ── 차트 생성 (시계열/집계값) ───────────────────────────
        report_dir = Path(settings.nfs_base_path) / "results" / job_id
        report_dir.mkdir(parents=True, exist_ok=True)
        chart_paths = _generate_all_charts(
            job_id, enriched, job_data["product_profile"], report_dir
        )

        context = {
            **job_data,
            "overall": overall,
            "fail_reasons": fail_reasons,
            "warn_reasons": warn_reasons,
            "summary": summary,
            "generated_at": generated_at,
            "check_results": enriched,
            "preflight_results": _by_phase("preflight"),
            "post_install_results": _by_phase("post_install"),
            "collect_results": _by_phase("collect"),
            "unknown_results": _by_phase("unknown"),
            "pass_count": all_statuses.count("pass"),
            "warn_count": all_statuses.count("warn"),
            "fail_count": all_statuses.count("fail"),
            "chart_gpu_path": chart_paths.get("gpu"),
            "chart_gpu_cooling_path": chart_paths.get("gpu_cooling"),
            "chart_cpu_path": chart_paths.get("cpu"),
            "chart_nccl_path": chart_paths.get("nccl"),
        }

        pdf_path = report_dir / "report.pdf"
        xlsx_path = report_dir / "report.xlsx"

        # ── 5. PDF + XLSX 생성 (동기 I/O — WeasyPrint 제약) ──
        log.info("report.render_start", job_id=job_id, overall=overall)
        _render_pdf(context, pdf_path)
        _render_xlsx(context, xlsx_path)
        log.info("report.render_done", job_id=job_id, pdf=str(pdf_path), xlsx=str(xlsx_path))

        # ── 6. DB에 Report 레코드 저장 ────────────────────────
        async with SessionLocal() as session:
            await _save_report_record(session, job_id, str(pdf_path), str(xlsx_path))

        # ── 7. Job.status → overall 결과로 확정 ──────────────
        _verdict_to_status = {"pass": "pass", "fail": "failed", "rejected": "rejected"}
        final_status = _verdict_to_status.get(overall, "failed")
        async with SessionLocal() as session:
            await _update_job_status(session, job_id, final_status)
        await publish_job_status(job_id, final_status)

        log.info("report.complete", job_id=job_id)
    finally:
        await engine.dispose()


async def _mark_error(job_id: str, message: str) -> None:
    engine, SessionLocal = _make_session()
    try:
        async with SessionLocal() as session:
            await _update_job_status(session, job_id, "error", message[:2000])
        await publish_job_status(job_id, "error", message[:2000])
    finally:
        await engine.dispose()


# ---------------------------------------------------------------------------
# Celery Task
# ---------------------------------------------------------------------------


@app.task(
    bind=True,
    queue="q_report",
    acks_late=True,
    max_retries=3,
    default_retry_delay=30,
    name="workers.report.generate_report",
)
def generate_report(self, job_id: str) -> dict:
    """
    PDF + XLSX 리포트 생성 태스크.

    Args:
        job_id: Job UUID (str)
    """
    log.info("report.start", job_id=job_id)
    try:
        asyncio.run(_async_generate_report(job_id))
        return {"job_id": job_id, "result": "ok"}
    except FileNotFoundError as exc:
        asyncio.run(_mark_error(job_id, str(exc)))
        raise self.retry(exc=exc)
    except Exception as exc:
        asyncio.run(_mark_error(job_id, str(exc)))
        raise self.retry(exc=exc)
