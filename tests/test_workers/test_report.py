"""
Report Worker 유닛 테스트.
DB·WeasyPrint·openpyxl은 mock — 렌더링 컨텍스트 구성, 상태 전이, 에러 경로 검증.
"""

import json
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# _render_xlsx — 실제 openpyxl 호출 (파일 I/O 없이 메모리에서 검증)
# ---------------------------------------------------------------------------


def _make_context(overall: str = "pass") -> dict:
    check_results = [
        {
            "check_name": "sw_gpu_hw",
            "display_name": "GPU 하드웨어 (PCIe)",
            "status": "pass",
            "detail": "gpu_count=4|link_width=16|link_speed=16GT/s",
            "claude_verdict": "[PASS] 정상",
            "phase": "preflight",
            "display_fields": [("gpu_count", "4"), ("link_width", "16"), ("link_speed", "16GT/s")],
        },
        {
            "check_name": "sw_power_mgmt",
            "display_name": "전원 관리",
            "status": "fail" if overall == "fail" else "pass",
            "detail": "sleep_target=active|cpu_governor=performance",
            "claude_verdict": "[FAIL] 미마스킹" if overall == "fail" else "[PASS] 정상",
            "phase": "preflight",
            "display_fields": [("sleep_target", "active"), ("cpu_governor", "performance")],
        },
    ]
    return {
        "job_id": "aaaaaaaa-0000-0000-0000-000000000001",
        "target_host": "10.0.0.1",
        "target_user": "root",
        "product_profile": "gpu_server",
        "created_at": "2026-03-25 12:00:00 UTC",
        "generated_at": "2026-03-25 12:05:00 UTC",
        "overall": overall,
        "fail_reasons": ["GPU 온도 92°C > 87°C"] if overall == "fail" else [],
        "warn_reasons": [],
        "summary": "테스트 요약",
        "check_results": check_results,
        "preflight_results": check_results,
        "post_install_results": [],
        "collect_results": [],
        "unknown_results": [],
        "pass_count": 2 if overall == "pass" else 1,
        "warn_count": 0,
        "fail_count": 1 if overall == "fail" else 0,
    }


def test_render_xlsx_sheets(tmp_path):
    from workers.report import _render_xlsx

    out = tmp_path / "report.xlsx"
    _render_xlsx(_make_context("pass"), out)

    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    assert "요약" in wb.sheetnames
    assert "검수 상세" in wb.sheetnames


def test_render_xlsx_summary_verdict(tmp_path):
    from workers.report import _render_xlsx

    out = tmp_path / "report.xlsx"
    _render_xlsx(_make_context("fail"), out)

    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    ws = wb["요약"]
    values = [ws.cell(row=r, column=2).value for r in range(1, 8)]
    assert "FAIL" in values


def test_render_xlsx_detail_rows(tmp_path):
    from workers.report import _render_xlsx

    out = tmp_path / "report.xlsx"
    ctx = _make_context("pass")
    _render_xlsx(ctx, out)

    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    ws = wb["검수 상세"]
    # 헤더(1) + 검수 항목(2)
    assert ws.max_row == 3
    # 1열=display_name, 2열=script(check_name)
    assert ws.cell(row=2, column=1).value == "GPU 하드웨어 (PCIe)"
    assert ws.cell(row=2, column=2).value == "sw_gpu_hw"


def test_render_xlsx_fail_reasons(tmp_path):
    from workers.report import _render_xlsx

    out = tmp_path / "report.xlsx"
    _render_xlsx(_make_context("fail"), out)

    import openpyxl

    wb = openpyxl.load_workbook(str(out))
    ws = wb["요약"]
    all_values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "GPU 온도 92°C > 87°C" in all_values


# ---------------------------------------------------------------------------
# _render_pdf — xelatex 호출 검증 (실제 컴파일 없이 mock)
# ---------------------------------------------------------------------------


def test_render_pdf_calls_xelatex(tmp_path):
    from workers.report import _render_pdf

    out = tmp_path / "report.pdf"
    with (
        patch("workers.report.subprocess.run") as mock_run,
        patch("workers.report.shutil.copy"),
    ):
        mock_run.return_value = MagicMock(returncode=0, stdout="")
        _render_pdf(_make_context("pass"), out)

    mock_run.assert_called_once()
    cmd = mock_run.call_args[0][0]
    assert cmd[0] == "xelatex"
    assert "-interaction=nonstopmode" in cmd


def test_render_pdf_xelatex_failure(tmp_path):
    """xelatex 실패 시 RuntimeError 발생."""
    from workers.report import _render_pdf

    out = tmp_path / "report.pdf"
    with patch("workers.report.subprocess.run") as mock_run:
        mock_run.return_value = MagicMock(returncode=1, stdout="! LaTeX Error: something")
        with pytest.raises(RuntimeError, match="xelatex failed"):
            _render_pdf(_make_context("pass"), out)


def test_render_pdf_template_renders_without_error(tmp_path):
    """pass/fail 컨텍스트가 LaTeX 템플릿에 정상 렌더링되는지 확인 (컴파일 없이)."""
    from workers.report import _latex_env

    for overall in ("pass", "fail", "error"):
        tex_str = _latex_env.get_template("report.tex.j2").render(**_make_context(overall))
        assert r"\documentclass" in tex_str
        assert overall.upper() in tex_str.upper() or "ERROR" in tex_str


# ---------------------------------------------------------------------------
# _async_generate_report — DB·NFS·렌더러 모두 mock하여 흐름 검증
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_async_generate_report_pass_flow(tmp_path):
    """pass 흐름: verdict 로드 → 렌더링 → DB 저장 → status=pass."""
    job_id = str(uuid.uuid4())
    verdict = {
        "verdict": "pass",
        "fail_items": [],
        "warn_items": [],
        "warn_count": 0,
        "agent_verdict": None,
    }

    # NFS verdict 파일 생성
    verdict_dir = tmp_path / "results" / job_id
    verdict_dir.mkdir(parents=True)
    (verdict_dir / "claude_verdict.json").write_text(json.dumps(verdict))

    fake_job = MagicMock()
    fake_job.id = uuid.UUID(job_id)
    fake_job.target_host = "10.0.0.1"
    fake_job.target_user = "root"
    fake_job.product_profile = "gpu_server"
    fake_job.created_at = MagicMock()
    fake_job.created_at.strftime.return_value = "2026-03-25 12:00:00 UTC"

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock) as mock_save,
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf") as mock_pdf,
        patch("workers.report._render_xlsx") as mock_xlsx,
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (fake_job, [])

        from workers.report import _async_generate_report

        await _async_generate_report(job_id)

    mock_pdf.assert_called_once()
    mock_xlsx.assert_called_once()
    mock_save.assert_called_once()

    # 마지막 update 호출이 status="pass"인지 확인
    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "pass"


@pytest.mark.asyncio
async def test_async_generate_report_missing_verdict_uses_fallback(tmp_path):
    """claude_verdict.json 없으면 DB check_results에서 fallback verdict 합성 후 리포트 생성."""
    job_id = str(uuid.uuid4())

    fake_job = MagicMock()
    fake_job.id = uuid.UUID(job_id)
    fake_job.target_host = "10.0.0.1"
    fake_job.target_user = "root"
    fake_job.product_profile = "gpu_server"
    fake_job.created_at = MagicMock()
    fake_job.created_at.strftime.return_value = "2026-03-25 12:00:00 UTC"

    fail_result = MagicMock()
    fail_result.check_name = "sw_gpu_hw"
    fail_result.status = "fail"
    fail_result.detail = "gpu_count=0"
    fail_result.claude_verdict = None

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock),
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf"),
        patch("workers.report._render_xlsx"),
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (fake_job, [fail_result])

        from workers.report import _async_generate_report

        # verdict 파일 없어도 예외 없이 완료
        await _async_generate_report(job_id)

    # fallback overall은 "fail" → 최종 status="failed"
    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "failed"


@pytest.mark.asyncio
async def test_async_generate_report_fail_verdict_sets_status_failed(tmp_path):
    """P0 회귀: verdict='fail' → job.status='failed' (deprecated 'fail' 아님)."""
    job_id = str(uuid.uuid4())
    verdict = {
        "verdict": "fail",
        "fail_items": [{"check": "sw_gpu_sw", "metric": "gpu_max_temp_c", "value": "95"}],
        "warn_items": [],
        "warn_count": 0,
        "agent_verdict": None,
    }

    verdict_dir = tmp_path / "results" / job_id
    verdict_dir.mkdir(parents=True)
    (verdict_dir / "claude_verdict.json").write_text(json.dumps(verdict))

    fake_job = MagicMock()
    fake_job.id = uuid.UUID(job_id)
    fake_job.target_host = "10.0.0.1"
    fake_job.target_user = "root"
    fake_job.product_profile = "gpu_server"
    fake_job.created_at = MagicMock()
    fake_job.created_at.strftime.return_value = "2026-03-25 12:00:00 UTC"

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock),
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf"),
        patch("workers.report._render_xlsx"),
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (fake_job, [])

        from workers.report import _async_generate_report

        await _async_generate_report(job_id)

    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "failed"


@pytest.mark.asyncio
async def test_async_generate_report_rejected_verdict_sets_status_rejected(tmp_path):
    """P0 회귀: verdict='rejected' → job.status='rejected'."""
    job_id = str(uuid.uuid4())
    verdict = {
        "verdict": "rejected",
        "fail_items": [],
        "warn_items": [{"check": "sw_gpu_sw", "metric": "gpu_max_temp_c", "value": "80"}],
        "warn_count": 1,
        "agent_verdict": {"verdict": "reject", "reason": "복합 경계값"},
    }

    verdict_dir = tmp_path / "results" / job_id
    verdict_dir.mkdir(parents=True)
    (verdict_dir / "claude_verdict.json").write_text(json.dumps(verdict))

    fake_job = MagicMock()
    fake_job.id = uuid.UUID(job_id)
    fake_job.target_host = "10.0.0.1"
    fake_job.target_user = "root"
    fake_job.product_profile = "gpu_server"
    fake_job.created_at = MagicMock()
    fake_job.created_at.strftime.return_value = "2026-03-25 12:00:00 UTC"

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock),
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf"),
        patch("workers.report._render_xlsx"),
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (fake_job, [])

        from workers.report import _async_generate_report

        await _async_generate_report(job_id)

    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "rejected"


def _fake_job_fixture(job_id: str) -> MagicMock:
    fake_job = MagicMock()
    fake_job.id = uuid.UUID(job_id)
    fake_job.target_host = "10.0.0.1"
    fake_job.target_user = "root"
    fake_job.product_profile = "gpu_server"
    fake_job.created_at = MagicMock()
    fake_job.created_at.strftime.return_value = "2026-03-25 12:00:00 UTC"
    return fake_job


@pytest.mark.asyncio
async def test_fallback_warn_only_yields_failed_not_pass(tmp_path):
    """회귀: verdict 없을 때 warn-only check_results → overall='warn' → status='failed' (pass 아님)."""
    job_id = str(uuid.uuid4())

    warn_result = MagicMock()
    warn_result.check_name = "sw_gpu_sw"
    warn_result.status = "warn"
    warn_result.detail = "gpu_max_temp_c=80"
    warn_result.claude_verdict = None

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock),
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf"),
        patch("workers.report._render_xlsx"),
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (_fake_job_fixture(job_id), [warn_result])

        from workers.report import _async_generate_report

        await _async_generate_report(job_id)

    # warn-only → overall="warn" → _verdict_to_status 미포함 → default "failed"
    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "failed"


@pytest.mark.asyncio
async def test_fallback_empty_check_results_yields_failed_not_pass(tmp_path):
    """회귀: verdict 없고 check_results 비어있을 때 overall='fail' → status='failed' (pass 아님).

    preflight 실패로 스크립트가 하나도 안 돌았을 때의 케이스.
    """
    job_id = str(uuid.uuid4())

    with (
        patch("workers.report.settings") as mock_settings,
        patch("workers.report._load_job_and_results", new_callable=AsyncMock) as mock_load,
        patch("workers.report._save_report_record", new_callable=AsyncMock),
        patch("workers.report._update_job_status", new_callable=AsyncMock) as mock_update,
        patch("workers.report.publish_job_status", new_callable=AsyncMock),
        patch("workers.report._render_pdf"),
        patch("workers.report._render_xlsx"),
        patch("workers.report._make_session", return_value=(AsyncMock(), MagicMock())),
    ):
        mock_settings.nfs_base_path = str(tmp_path)
        mock_load.return_value = (_fake_job_fixture(job_id), [])

        from workers.report import _async_generate_report

        await _async_generate_report(job_id)

    last_call_args = mock_update.call_args_list[-1]
    assert last_call_args.args[2] == "failed"
